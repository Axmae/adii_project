import io
import urllib.parse

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import F, Q, Count
from django.http import HttpResponse
from django.template.loader import render_to_string
from .models import Measurement, RetourEffet
from .forms import MeasurementForm, AdminNoteForm, RetourEffetForm
from notifications.utils import create_notification
from stock.models import StockItem, StockMovement

def role_required(roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('auth')
            if request.user.role not in roles:
                messages.error(request, "Accès non autorisé.")
                return redirect('home')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator

# ─── AGENT ───────────────────────────────────────────────
@login_required
def agent_dashboard(request):
    measurements = Measurement.objects.filter(user=request.user)
    retours = RetourEffet.objects.filter(agent=request.user)
    return render(request, 'agent/dashboard.html', {
        'measurements': measurements,
        'retours': retours,
    })

@login_required
def create_measurement(request):
    if request.user.role != 'agent':
        return redirect('home')
    existing = Measurement.objects.filter(
        user=request.user,
        status__in=['en_attente', 'valide', 'en_production', 'pret']
    ).exists()
    if existing:
        messages.warning(request, "Vous avez déjà une fiche en cours de traitement.")
        return redirect('agent_dashboard')
    if request.method == 'POST':
        form = MeasurementForm(request.POST)
        if form.is_valid():
            m = form.save(commit=False)
            m.user = request.user
            m.rempli_par = request.user
            m.save()
            from accounts.models import User
            for admin in User.objects.filter(role='admin'):
                create_notification(
                    admin,
                    'Nouvelle fiche de mesure',
                    f"{request.user.get_full_name()} a soumis une fiche pour {m.get_type_equipement_display()}.",
                    category='measurement'
                )
            messages.success(request, "Fiche créée avec succès !")
            return redirect('agent_dashboard')
    else:
        form = MeasurementForm()
    return render(request, 'agent/create_measurement.html', {'form': form})

@login_required
def edit_measurement(request, pk):
    m = get_object_or_404(Measurement, pk=pk, user=request.user)
    if not m.can_edit():
        messages.error(request, "Cette fiche ne peut plus être modifiée.")
        return redirect('agent_dashboard')
    if request.method == 'POST':
        form = MeasurementForm(request.POST, instance=m)
        if form.is_valid():
            form.save()
            messages.success(request, "Fiche mise à jour.")
            return redirect('agent_dashboard')
    else:
        form = MeasurementForm(instance=m)
    return render(request, 'agent/create_measurement.html', {'form': form, 'edit': True})

# ─── SECRÉTAIRE ──────────────────────────────────────────
@role_required(['secretaire', 'admin'])
def secretaire_dashboard(request):
    measurements = Measurement.objects.select_related('user', 'rempli_par').order_by('-created_at')
    return render(request, 'secretaire/dashboard.html', {
        'measurements': measurements,
    })

@role_required(['secretaire', 'admin'])
def secretaire_create(request):
    from accounts.models import User
    agents = User.objects.filter(role='agent').order_by('nom')
    if request.method == 'POST':
        agent_id = request.POST.get('agent_id')
        agent = get_object_or_404(User, pk=agent_id, role='agent')
        existing = Measurement.objects.filter(
            user=agent,
            status__in=['en_attente', 'valide', 'en_production', 'pret']
        ).exists()
        if existing:
            messages.warning(request, f"{agent.get_full_name()} a déjà une fiche en cours.")
            return redirect('secretaire_dashboard')
        form = MeasurementForm(request.POST)
        if form.is_valid():
            m = form.save(commit=False)
            m.user = agent
            m.rempli_par = request.user
            m.save()
            for admin in User.objects.filter(role='admin'):
                create_notification(
                    admin,
                    'Nouvelle fiche de mesure',
                    f"Fiche de {agent.get_full_name()} remplie par {request.user.get_full_name()}.",
                    category='measurement'
                )
            messages.success(request, f"Fiche créée pour {agent.get_full_name()}.")
            return redirect('secretaire_dashboard')
    else:
        form = MeasurementForm()
    selected_agent = None
    agent_id = request.GET.get('agent')
    if agent_id:
        selected_agent = User.objects.filter(pk=agent_id, role='agent').first()
    return render(request, 'secretaire/create.html', {
        'form': form,
        'agents': agents,
        'selected_agent': selected_agent,
    })

# ─── ADMIN ───────────────────────────────────────────────
@role_required(['admin'])
def admin_dashboard(request):
    from accounts.models import User
    measurements = Measurement.objects.select_related('user').all()
    stats = {
        'total':         measurements.count(),
        'en_attente':    measurements.filter(status='en_attente').count(),
        'en_production': measurements.filter(status='en_production').count(),
        'pret':          measurements.filter(status='pret').count(),
        'livre':         measurements.filter(status='livre').count(),
        'agents':        User.objects.filter(role='agent').count(),
    }
    stock_alerts = StockItem.objects.filter(
        quantity__lt=F('min_threshold')
    ).order_by('quantity')
    recent_stock_movements = StockMovement.objects.select_related(
        'stock_item', 'created_by'
    ).all()[:8]
    return render(request, 'admin_panel/dashboard.html', {
        'measurements':           measurements,
        'stats':                  stats,
        'stock_alerts':           stock_alerts,
        'recent_stock_movements': recent_stock_movements,
    })

@role_required(['admin'])
def validate_measurement(request, pk):
    m = get_object_or_404(Measurement, pk=pk)
    action = request.POST.get('action')
    if action == 'valide':
        m.status = 'valide'
        m.save()
        from accounts.models import User
        for tech in User.objects.filter(role='technicien'):
            create_notification(
                tech,
                'Nouvelle fiche à traiter',
                f"La fiche de {m.user.get_full_name()} a été validée.",
                category='measurement'
            )
        create_notification(
            m.user,
            'Fiche validée',
            f"Votre demande pour {m.get_type_equipement_display()} a été validée.",
            category='measurement'
        )
        messages.success(request, "Fiche validée.")
    elif action == 'refuse':
        m.status = 'refuse'
        m.notes_admin = request.POST.get('notes_admin', '')
        m.save()
        create_notification(
            m.user,
            'Fiche refusée',
            f"Votre demande a été refusée. {m.notes_admin}",
            category='measurement'
        )
        messages.warning(request, "Fiche refusée.")
    return redirect('admin_dashboard')

@role_required(['admin'])
def manage_users(request):
    from accounts.models import User
    users = User.objects.all().order_by('role', 'nom')
    return render(request, 'admin_panel/users.html', {'users': users})

@role_required(['admin'])
def change_role(request, user_id):
    from accounts.models import User
    user = get_object_or_404(User, pk=user_id)
    new_role = request.POST.get('role')
    if new_role in ['agent', 'admin', 'technicien', 'secretaire']:
        user.role = new_role
        user.save()
        messages.success(request, f"Rôle de {user.get_full_name()} mis à jour.")
    return redirect('manage_users')

@role_required(['admin'])
def livraison_groupee(request):
    fiches_pretes = Measurement.objects.filter(
        status='pret'
    ).select_related('user').order_by('user__nom')
    return render(request, 'admin_panel/livraison_groupee.html', {
        'fiches_pretes': fiches_pretes,
    })

@role_required(['admin'])
def confirmer_livraison_groupee(request):
    if request.method == 'POST':
        ids = request.POST.getlist('fiche_ids')
        if not ids:
            messages.warning(request, "Aucune fiche sélectionnée.")
            return redirect('livraison_groupee')
        fiches = Measurement.objects.filter(pk__in=ids, status='pret')
        count = fiches.count()
        for fiche in fiches:
            fiche.status = 'livre'
            fiche.save()
            create_notification(
                fiche.user,
                'Tenue livrée ✅',
                f"Votre tenue ({fiche.get_type_equipement_display()}) vous a été remise officiellement.",
                category='measurement'
            )
        messages.success(request, f"{count} tenue(s) marquée(s) comme livrée(s).")
        return redirect('livraison_groupee')
    return redirect('livraison_groupee')

@role_required(['admin', 'secretaire'])
def avancement_groupe(request):
    from accounts.models import User
    fiches = Measurement.objects.filter(
        status__in=['en_attente', 'valide', 'en_production', 'pret']
    ).select_related('user', 'rempli_par').order_by('user__nom')
    agents = User.objects.filter(role='agent').order_by('nom')
    agent_filter  = request.GET.get('agent', '')
    status_filter = request.GET.get('status', '')
    if agent_filter:
        fiches = fiches.filter(user__pk=agent_filter)
    if status_filter:
        fiches = fiches.filter(status=status_filter)
    return render(request, 'admin_panel/avancement_groupe.html', {
        'fiches':         fiches,
        'agents':         agents,
        'agent_filter':   agent_filter,
        'status_filter':  status_filter,
        'status_choices': Measurement.STATUS_CHOICES,
    })

@role_required(['admin', 'secretaire'])
def confirmer_avancement_groupe(request):
    if request.method == 'POST':
        ids        = request.POST.getlist('fiche_ids')
        new_status = request.POST.get('new_status')
        valid_statuses = ['en_attente', 'valide', 'en_production', 'pret', 'livre', 'refuse']
        if not ids:
            messages.warning(request, "Aucune fiche sélectionnée.")
            return redirect('avancement_groupe')
        if new_status not in valid_statuses:
            messages.error(request, "Statut invalide.")
            return redirect('avancement_groupe')
        fiches = Measurement.objects.filter(pk__in=ids)
        count  = fiches.count()
        status_labels = {
            'en_attente':    'En attente',
            'valide':        'Validé',
            'en_production': 'En production',
            'pret':          'Prêt',
            'livre':         'Livré',
            'refuse':        'Refusé',
        }
        notif_msgs = {
            'valide':        "Votre fiche a été validée par le responsable.",
            'en_production': "Votre tenue est en cours de fabrication.",
            'pret':          "Votre tenue est prête ! Venez la récupérer.",
            'livre':         "Votre tenue vous a été remise officiellement.",
            'refuse':        "Votre demande a été refusée.",
            'en_attente':    "Votre fiche a été remise en attente.",
        }
        for fiche in fiches:
            fiche.status = new_status
            fiche.save()
            create_notification(
                fiche.user,
                f"Avancement mis à jour — {status_labels[new_status]}",
                notif_msgs.get(new_status, "Votre dossier a été mis à jour."),
                category='measurement'
            )
            if new_status == 'en_production':
                item = StockItem.objects.filter(category=fiche.type_equipement).first()
                if item and item.quantity > 0:
                    qty_before = item.quantity
                    item.quantity -= 1
                    item.save()
                    StockMovement.objects.create(
                        stock_item=item,
                        type='sortie',
                        reason='production',
                        quantity=1,
                        quantity_before=qty_before,
                        quantity_after=item.quantity,
                        note=f"Mise en production groupée — fiche #{fiche.pk} ({fiche.user.get_full_name()})",
                        created_by=request.user,
                    )
                    if item.quantity < item.min_threshold:
                        from accounts.models import User
                        for admin in User.objects.filter(role='admin'):
                            create_notification(
                                admin,
                                '⚠️ Stock bas',
                                f"{item.name} ({item.size}) : {item.quantity} unités restantes.",
                                category='stock'
                            )
        messages.success(request, f"{count} fiche(s) mise(s) à jour → {status_labels[new_status]}.")
        return redirect('avancement_groupe')
    return redirect('avancement_groupe')

# ─── TECHNICIEN ──────────────────────────────────────────
@role_required(['technicien'])
def tech_dashboard(request):
    measurements = Measurement.objects.filter(
        status__in=['valide', 'en_production', 'pret']
    ).select_related('user')
    return render(request, 'technicien/dashboard.html', {'measurements': measurements})

@role_required(['technicien'])
def update_status(request, pk):
    m = get_object_or_404(Measurement, pk=pk)
    new_status = request.POST.get('status')
    allowed = {
        'valide':        'en_production',
        'en_production': 'pret',
        'pret':          'livre',
    }
    if new_status == allowed.get(m.status):
        m.status = new_status
        m.save()
        if new_status == 'en_production':
            item = StockItem.objects.filter(category=m.type_equipement).first()
            if item and item.quantity > 0:
                qty_before = item.quantity
                item.quantity -= 1
                item.save()
                StockMovement.objects.create(
                    stock_item=item,
                    type='sortie',
                    reason='production',
                    quantity=1,
                    quantity_before=qty_before,
                    quantity_after=item.quantity,
                    note=f"Mise en production — fiche #{m.pk} ({m.user.get_full_name()})",
                    created_by=request.user,
                )
                if item.quantity < item.min_threshold:
                    from accounts.models import User
                    for admin in User.objects.filter(role='admin'):
                        create_notification(
                            admin,
                            '⚠️ Stock bas',
                            f"{item.name} ({item.size}) : {item.quantity} unités restantes (seuil : {item.min_threshold}).",
                            category='stock'
                        )
        msgs = {
            'en_production': "Votre équipement est en cours de fabrication.",
            'pret':          "Votre équipement est prêt ! Venez le récupérer.",
            'livre':         "Votre équipement vous a été remis.",
        }
        if new_status in msgs:
            create_notification(
                m.user,
                "Statut mis à jour",
                msgs[new_status],
                category='measurement'
            )
        messages.success(request, f"Statut mis à jour : {m.get_status_display()}")
    return redirect('tech_dashboard')

# ─── RETOUR D'EFFETS (Agent) ──────────────────────────────
@login_required
def enregistrer_retour(request):
    if request.user.role != 'agent':
        return redirect('home')
    if request.method == 'POST':
        form = RetourEffetForm(request.POST)
        if form.is_valid():
            r = form.save(commit=False)
            r.agent = request.user
            r.created_by = request.user
            r.save()
            from accounts.models import User
            for admin in User.objects.filter(role='admin'):
                create_notification(
                    admin,
                    'Retour d\'effet déclaré',
                    f"{request.user.get_full_name()} a retourné {r.quantite} x {r.get_type_equipement_display()} pour {r.get_motif_display()}.",
                    category='stock'
                )
            messages.success(request, "Retour enregistré avec succès.")
            return redirect('agent_dashboard')
    else:
        form = RetourEffetForm()
    return render(request, 'agent/retour.html', {'form': form})


# ─── HISTORIQUE PAR AGENT (Admin) ─────────────────────────
@role_required(['admin'])
def historique_agents(request):
    from accounts.models import User
    agents = User.objects.filter(role='agent').order_by('nom')
    agent_filter = request.GET.get('agent', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    if agent_filter:
        agents = agents.filter(pk=agent_filter)

    historiques = []
    for agent in agents:
        measurements = Measurement.objects.filter(user=agent)
        retours = RetourEffet.objects.filter(agent=agent)
        if date_debut:
            measurements = measurements.filter(created_at__date__gte=date_debut)
            retours = retours.filter(created_at__date__gte=date_debut)
        if date_fin:
            measurements = measurements.filter(created_at__date__lte=date_fin)
            retours = retours.filter(created_at__date__lte=date_fin)
        historiques.append({
            'agent': agent,
            'measurements': measurements,
            'retours': retours,
        })

    return render(request, 'admin_panel/historique.html', {
        'historiques': historiques,
        'agents': User.objects.filter(role='agent').order_by('nom'),
        'agent_filter': agent_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
    })


# ─── EXPORT EXCEL ─────────────────────────────────────────
@role_required(['admin'])
def export_historique_excel(request):
    from accounts.models import User
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historique par agent"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = [
        "Agent", "Matricule", "Service",
        "Date demande", "Équipement demandé", "Statut demande",
        "Date retour", "Équipement retourné", "Quantité retour", "Motif retour"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    row = 2
    agent_filter = request.GET.get('agent', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    agents_qs = User.objects.filter(role='agent').order_by('nom')
    if agent_filter:
        agents_qs = agents_qs.filter(pk=agent_filter)

    for agent in agents_qs:
        measurements = Measurement.objects.filter(user=agent)
        retours = RetourEffet.objects.filter(agent=agent)
        if date_debut:
            measurements = measurements.filter(created_at__date__gte=date_debut)
            retours = retours.filter(created_at__date__gte=date_debut)
        if date_fin:
            measurements = measurements.filter(created_at__date__lte=date_fin)
            retours = retours.filter(created_at__date__lte=date_fin)

        measurements = list(measurements)
        retours = list(retours)
        max_rows = max(len(measurements), len(retours), 1)
        for i in range(max_rows):
            m = measurements[i] if i < len(measurements) else None
            r = retours[i] if i < len(retours) else None
            ws.cell(row=row, column=1, value=agent.get_full_name()).border = thin_border
            ws.cell(row=row, column=2, value=agent.matricule or '').border = thin_border
            ws.cell(row=row, column=3, value=agent.service or '').border = thin_border
            if m:
                ws.cell(row=row, column=4, value=m.created_at.strftime('%d/%m/%Y')).border = thin_border
                ws.cell(row=row, column=5, value=m.get_type_equipement_display()).border = thin_border
                ws.cell(row=row, column=6, value=m.get_status_display()).border = thin_border
            if r:
                ws.cell(row=row, column=7, value=r.created_at.strftime('%d/%m/%Y')).border = thin_border
                ws.cell(row=row, column=8, value=r.get_type_equipement_display()).border = thin_border
                ws.cell(row=row, column=9, value=r.quantite).border = thin_border
                ws.cell(row=row, column=10, value=r.get_motif_display()).border = thin_border
            row += 1

    for col in range(1, 11):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "historique_agents.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── EXPORT PDF ───────────────────────────────────────────
def _render_pdf_reportlab(historiques, agents, agent_filter, date_debut, date_fin):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), title="Historique par agent")
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle('Title2', parent=styles['Title'], fontSize=16, spaceAfter=6)
    elements.append(Paragraph("ADII — Historique par agent", title_style))
    subtitle = f"Filtres: {'Agent spécifique' if agent_filter else 'Tous les agents'}"
    if date_debut or date_fin:
        subtitle += f" | Du {date_debut or '…'} au {date_fin or '…'}"
    elements.append(Paragraph(subtitle, styles['Normal']))
    elements.append(Spacer(1, 12))

    for h in historiques:
        agent = h['agent']
        measurements = list(h['measurements'])
        retours = list(h['retours'])
        if not measurements and not retours:
            continue

        elements.append(Paragraph(
            f"<b>{agent.get_full_name()}</b> — {agent.matricule or 'N/A'} — {agent.service or 'N/A'}",
            styles['Heading2']
        ))

        if measurements:
            data = [["Date", "Équipement", "Statut"]]
            for m in measurements:
                data.append([
                    m.created_at.strftime('%d/%m/%Y'),
                    m.get_type_equipement_display(),
                    m.get_status_display()
                ])
            table = Table(data, colWidths=[80, 150, 120])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A3A6B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F0F4F8')]),
            ]))
            elements.append(Paragraph("<b>Demandes:</b>", styles['Normal']))
            elements.append(table)
            elements.append(Spacer(1, 8))

        if retours:
            data = [["Date", "Équipement", "Qté", "Motif"]]
            for r in retours:
                data.append([
                    r.created_at.strftime('%d/%m/%Y'),
                    r.get_type_equipement_display(),
                    str(r.quantite),
                    r.get_motif_display()
                ])
            table = Table(data, colWidths=[80, 150, 60, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#991B1B')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FEF2F2')]),
            ]))
            elements.append(Paragraph("<b>Retours:</b>", styles['Normal']))
            elements.append(table)
            elements.append(Spacer(1, 8))

        elements.append(Spacer(1, 12))

    doc.build(elements)
    buffer.seek(0)
    return buffer


@role_required(['admin'])
def export_historique_pdf(request):
    from accounts.models import User
    agent_filter = request.GET.get('agent', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    agents_qs = User.objects.filter(role='agent').order_by('nom')
    if agent_filter:
        agents_qs = agents_qs.filter(pk=agent_filter)

    historiques = []
    for agent in agents_qs:
        measurements = Measurement.objects.filter(user=agent)
        retours = RetourEffet.objects.filter(agent=agent)
        if date_debut:
            measurements = measurements.filter(created_at__date__gte=date_debut)
            retours = retours.filter(created_at__date__gte=date_debut)
        if date_fin:
            measurements = measurements.filter(created_at__date__lte=date_fin)
            retours = retours.filter(created_at__date__lte=date_fin)
        historiques.append({
            'agent': agent,
            'measurements': measurements,
            'retours': retours,
        })

    try:
        import weasyprint
        html = render_to_string('admin_panel/historique_pdf.html', {
            'historiques': historiques,
            'agent_filter': agent_filter,
            'date_debut': date_debut,
            'date_fin': date_fin,
        })
        pdf = weasyprint.HTML(string=html).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="historique_agents.pdf"'
        return response
    except (ImportError, OSError):
        buffer = _render_pdf_reportlab(historiques, agents_qs, agent_filter, date_debut, date_fin)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="historique_agents.pdf"'
        return response


@login_required
def measurement_detail(request, pk):
    m = get_object_or_404(Measurement, pk=pk)
    if request.user.role == 'agent' and m.user != request.user:
        return redirect('home')
    return render(request, 'measurement_detail.html', {'m': m})