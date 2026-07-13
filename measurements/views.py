import io
import urllib.parse

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import F, Q, Count, Sum
from django.http import HttpResponse, HttpResponseBadRequest
from django.template.loader import render_to_string
from .models import Measurement, RetourEffet
from .forms import MeasurementForm, AdminNoteForm, RetourEffetForm
from notifications.utils import create_notification
from notifications.email_utils import send_status_email
from stock.models import StockItem, StockMovement

def role_required(roles):
    def decorator(view_func):
        def wrapper(request, *args, **kwargs):
            if not request.user.is_authenticated:
                return redirect('auth')
            if request.user.role not in roles:
                messages.error(request, "Accès non autorisé.")
                return redirect('home')
            return view_func(request, *args, **kwargs)
        return wrapper
    return decorator

# ─── AGENT ───────────────────────────────────────────────
@login_required
def agent_dashboard(request):
    measurements = Measurement.objects.filter(user=request.user).order_by('-created_at')[:50]
    retours = RetourEffet.objects.filter(agent=request.user).order_by('-created_at')[:20]
    return render(request, 'agent/dashboard.html', {
        'measurements': measurements,
        'retours': retours,
    })

@login_required
def create_measurement(request):
    if request.user.role != 'agent':
        return redirect('home')
    existing = Measurement.objects.filter(
        user=request.user,
        status__in=['en_attente', 'valide', 'en_production', 'pret']
    ).exists()
    if existing:
        messages.warning(request, "Vous avez déjà une fiche en cours de traitement.")
        return redirect('agent_dashboard')
    if request.method == 'POST':
        form = MeasurementForm(request.POST)
        if form.is_valid():
            m = form.save(commit=False)
            m.user = request.user
            m.rempli_par = request.user
            m.save()
            send_status_email(m)
            from accounts.models import User
            for admin in User.objects.filter(role='admin'):
                create_notification(
                    admin,
                    'Nouvelle fiche de mesure',
                    f"{request.user.get_full_name()} a soumis une fiche pour {m.get_type_equipement_display()}.",
                    category='measurement'
                )
            messages.success(request, "Fiche créée avec succès !")
            return redirect('agent_dashboard')
    else:
        form = MeasurementForm()
    return render(request, 'agent/create_measurement.html', {'form': form})

@login_required
def edit_measurement(request, pk):
    m = get_object_or_404(Measurement, pk=pk, user=request.user)
    if not m.can_edit():
        messages.error(request, "Cette fiche ne peut plus être modifiée.")
        return redirect('agent_dashboard')
    if request.method == 'POST':
        form = MeasurementForm(request.POST, instance=m)
        if form.is_valid():
            form.save()
            messages.success(request, "Fiche mise à jour.")
            return redirect('agent_dashboard')
    else:
        form = MeasurementForm(instance=m)
    return render(request, 'agent/create_measurement.html', {'form': form, 'edit': True})

# ─── SECRÉTAIRE ──────────────────────────────────────────
@role_required(['secretaire', 'admin'])
def secretaire_dashboard(request):
    measurements = Measurement.objects.select_related('user', 'rempli_par').order_by('-created_at')[:200]
    return render(request, 'secretaire/dashboard.html', {
        'measurements': measurements,
    })

@role_required(['secretaire', 'admin'])
def secretaire_create(request):
    from accounts.models import User
    agents = User.objects.filter(role='agent').order_by('nom')
    if request.method == 'POST':
        agent_id = request.POST.get('agent_id')
        agent = get_object_or_404(User, pk=agent_id, role='agent')
        existing = Measurement.objects.filter(
            user=agent,
            status__in=['en_attente', 'valide', 'en_production', 'pret']
        ).exists()
        if existing:
            messages.warning(request, f"{agent.get_full_name()} a déjà une fiche en cours.")
            return redirect('secretaire_dashboard')
        form = MeasurementForm(request.POST)
        if form.is_valid():
            m = form.save(commit=False)
            m.user = agent
            m.rempli_par = request.user
            m.save()
            send_status_email(m)
            for admin in User.objects.filter(role='admin'):
                create_notification(
                    admin,
                    'Nouvelle fiche de mesure',
                    f"Fiche de {agent.get_full_name()} remplie par {request.user.get_full_name()}.",
                    category='measurement'
                )
            messages.success(request, f"Fiche créée pour {agent.get_full_name()}.")
            return redirect('secretaire_dashboard')
    else:
        form = MeasurementForm()
    selected_agent = None
    agent_id = request.GET.get('agent')
    if agent_id:
        selected_agent = User.objects.filter(pk=agent_id, role='agent').first()
    return render(request, 'secretaire/create.html', {
        'form': form,
        'agents': agents,
        'selected_agent': selected_agent,
    })

# ─── ADMIN ───────────────────────────────────────────────
@role_required(['admin'])
def admin_dashboard(request):
    from accounts.models import User
    stats_agg = Measurement.objects.aggregate(
        total=Count('id'),
        en_attente=Count('id', filter=Q(status='en_attente')),
        en_production=Count('id', filter=Q(status='en_production')),
        pret=Count('id', filter=Q(status='pret')),
        livre=Count('id', filter=Q(status='livre')),
    )
    stats_agg['agents'] = User.objects.filter(role='agent').count()
    measurements = Measurement.objects.select_related('user').order_by('-created_at')[:100]
    stock_alerts = StockItem.objects.filter(
        quantity__lt=F('min_threshold')
    ).order_by('quantity')
    recent_stock_movements = StockMovement.objects.select_related(
        'stock_item', 'created_by'
    ).order_by('-created_at')[:8]
    return render(request, 'admin_panel/dashboard.html', {
        'measurements':           measurements,
        'stats':                  stats_agg,
        'stock_alerts':           stock_alerts,
        'recent_stock_movements': recent_stock_movements,
    })

@role_required(['admin'])
def validate_measurement(request, pk):
    m = get_object_or_404(Measurement.objects.select_related('user'), pk=pk)
    action = request.POST.get('action')
    if action == 'valide':
        m.status = 'valide'
        m.save()
        send_status_email(m)
        from accounts.models import User
        for tech in User.objects.filter(role='technicien'):
            create_notification(
                tech,
                'Nouvelle fiche à traiter',
                f"La fiche de {m.user.get_full_name()} a été validée.",
                category='measurement'
            )
        create_notification(
            m.user,
            'Fiche validée',
            f"Votre demande pour {m.get_type_equipement_display()} a été validée.",
            category='measurement'
        )
        messages.success(request, "Fiche validée.")
    elif action == 'refuse':
        m.status = 'refuse'
        m.notes_admin = request.POST.get('notes_admin', '')
        m.save()
        send_status_email(m)
        create_notification(
            m.user,
            'Fiche refusée',
            f"Votre demande a été refusée. {m.notes_admin}",
            category='measurement'
        )
        messages.warning(request, "Fiche refusée.")
    return redirect('admin_dashboard')

@role_required(['admin'])
def manage_users(request):
    from accounts.models import User
    users = User.objects.all().order_by('role', 'nom')
    return render(request, 'admin_panel/users.html', {'users': users})

@role_required(['admin'])
def change_role(request, user_id):
    from accounts.models import User
    user = get_object_or_404(User, pk=user_id)
    new_role = request.POST.get('role')
    if new_role in ['agent', 'admin', 'technicien', 'secretaire']:
        user.role = new_role
        user.save()
        messages.success(request, f"Rôle de {user.get_full_name()} mis à jour.")
    return redirect('manage_users')

@role_required(['admin'])
def livraison_groupee(request):
    fiches_pretes = Measurement.objects.filter(
        status='pret'
    ).select_related('user').order_by('user__nom')
    last_confirmed = request.session.pop('last_confirmed_ids', None)
    return render(request, 'admin_panel/livraison_groupee.html', {
        'fiches_pretes': fiches_pretes,
        'last_confirmed_ids': ','.join(last_confirmed) if last_confirmed else '',
    })

@role_required(['admin'])
def confirmer_livraison_groupee(request):
    if request.method == 'POST':
        ids = request.POST.getlist('fiche_ids')
        if not ids:
            messages.warning(request, "Aucune fiche sélectionnée.")
            return redirect('livraison_groupee')
        fiches = list(Measurement.objects.filter(pk__in=ids, status='pret').select_related('user'))
        count = len(fiches)
        if count:
            from notifications.models import Notification
            for fiche in fiches:
                fiche.status = 'livre'
            Measurement.objects.bulk_update(fiches, ['status'])
            notifications = []
            for fiche in fiches:
                send_status_email(fiche)
                notifications.append(Notification(
                    user=fiche.user,
                    title='Tenue livrée',
                    message=f"Votre tenue ({fiche.get_type_equipement_display()}) vous a été remise officiellement.",
                    category='measurement'
                ))
            Notification.objects.bulk_create(notifications)
        messages.success(request, f"{count} tenue(s) marquée(s) comme livrée(s).")
        request.session['last_confirmed_ids'] = ids
        return redirect('livraison_groupee')
    return redirect('livraison_groupee')

@role_required(['admin', 'secretaire'])
def avancement_groupe(request):
    from accounts.models import User
    fiches = Measurement.objects.filter(
        status__in=['en_attente', 'valide', 'en_production', 'pret']
    ).select_related('user', 'rempli_par').order_by('user__nom')
    agents = User.objects.filter(role='agent').order_by('nom')
    agent_filter  = request.GET.get('agent', '')
    status_filter = request.GET.get('status', '')
    if agent_filter:
        fiches = fiches.filter(user__pk=agent_filter)
    if status_filter:
        fiches = fiches.filter(status=status_filter)
    return render(request, 'admin_panel/avancement_groupe.html', {
        'fiches':         fiches,
        'agents':         agents,
        'agent_filter':   agent_filter,
        'status_filter':  status_filter,
        'status_choices': Measurement.STATUS_CHOICES,
    })

@role_required(['admin', 'secretaire'])
def confirmer_avancement_groupe(request):
    if request.method == 'POST':
        ids        = request.POST.getlist('fiche_ids')
        new_status = request.POST.get('new_status')
        valid_statuses = ['en_attente', 'valide', 'en_production', 'pret', 'livre', 'refuse']
        if not ids:
            messages.warning(request, "Aucune fiche sélectionnée.")
            return redirect('avancement_groupe')
        if new_status not in valid_statuses:
            messages.error(request, "Statut invalide.")
            return redirect('avancement_groupe')
        fiches = list(Measurement.objects.filter(pk__in=ids).select_related('user'))
        count = len(fiches)
        if not count:
            messages.warning(request, "Aucune fiche trouvée.")
            return redirect('avancement_groupe')
        status_labels = {
            'en_attente':    'En attente',
            'valide':        'Validé',
            'en_production': 'En production',
            'pret':          'Prêt',
            'livre':         'Livré',
            'refuse':        'Refusé',
        }
        notif_msgs = {
            'valide':        "Votre fiche a été validée par le responsable.",
            'en_production': "Votre tenue est en cours de fabrication.",
            'pret':          "Votre tenue est prête ! Venez la récupérer.",
            'livre':         "Votre tenue vous a été remise officiellement.",
            'refuse':        "Votre demande a été refusée.",
            'en_attente':    "Votre fiche a été remise en attente.",
        }

        from accounts.models import User
        from notifications.models import Notification

        # Prefetch stock items for en_production case
        categories = set()
        for fiche in fiches:
            if new_status == 'en_production':
                categories.add(fiche.type_equipement)
        stock_map = {}
        if categories:
            for si in StockItem.objects.filter(category__in=categories):
                stock_map[si.category] = si

        # Bulk update statuses
        for fiche in fiches:
            fiche.status = new_status
        Measurement.objects.bulk_update(fiches, ['status'])

        notifications = []
        alerted_stock = set()

        for fiche in fiches:
            send_status_email(fiche)
            notifications.append(Notification(
                user=fiche.user,
                title=f"Avancement mis a jour — {status_labels[new_status]}",
                message=notif_msgs.get(new_status, "Votre dossier a été mis à jour."),
                category='measurement'
            ))
            if new_status == 'en_production':
                item = stock_map.get(fiche.type_equipement)
                if item and item.quantity > 0:
                    qty_before = item.quantity
                    item.quantity -= 1
                    item.save()
                    StockMovement.objects.create(
                        stock_item=item,
                        type='sortie',
                        reason='production',
                        quantity=1,
                        quantity_before=qty_before,
                        quantity_after=item.quantity,
                        note=f"Mise en production groupee — fiche #{fiche.pk} ({fiche.user.get_full_name()})",
                        created_by=request.user,
                    )
                    if item.quantity < item.min_threshold and item.pk not in alerted_stock:
                        alerted_stock.add(item.pk)
                        for admin in User.objects.filter(role='admin'):
                            notifications.append(Notification(
                                user=admin,
                                title='Stock bas',
                                message=f"{item.name} ({item.size}) : {item.quantity} unites restantes.",
                                category='stock'
                            ))

        Notification.objects.bulk_create(notifications)
        messages.success(request, f"{count} fiche(s) mise(s) à jour → {status_labels[new_status]}.")
        return redirect('avancement_groupe')
    return redirect('avancement_groupe')

# ─── TECHNICIEN ──────────────────────────────────────────
@role_required(['technicien'])
def tech_dashboard(request):
    measurements = Measurement.objects.filter(
        status__in=['valide', 'en_production', 'pret']
    ).select_related('user').order_by('-created_at')[:200]
    return render(request, 'technicien/dashboard.html', {'measurements': measurements})

@role_required(['technicien'])
def update_status(request, pk):
    m = get_object_or_404(Measurement.objects.select_related('user'), pk=pk)
    new_status = request.POST.get('status')
    allowed = {
        'valide':        'en_production',
        'en_production': 'pret',
        'pret':          'livre',
    }
    if new_status == allowed.get(m.status):
        m.status = new_status
        m.save()
        send_status_email(m)
        if new_status == 'en_production':
            item = StockItem.objects.filter(category=m.type_equipement).first()
            if item and item.quantity > 0:
                qty_before = item.quantity
                item.quantity -= 1
                item.save()
                StockMovement.objects.create(
                    stock_item=item,
                    type='sortie',
                    reason='production',
                    quantity=1,
                    quantity_before=qty_before,
                    quantity_after=item.quantity,
                    note=f"Mise en production — fiche #{m.pk} ({m.user.get_full_name()})",
                    created_by=request.user,
                )
                if item.quantity < item.min_threshold:
                    from accounts.models import User
                    for admin in User.objects.filter(role='admin'):
                        create_notification(
                            admin,
                            '⚠️ Stock bas',
                            f"{item.name} ({item.size}) : {item.quantity} unités restantes (seuil : {item.min_threshold}).",
                            category='stock'
                        )
        msgs = {
            'en_production': "Votre équipement est en cours de fabrication.",
            'pret':          "Votre équipement est prêt ! Venez le récupérer.",
            'livre':         "Votre équipement vous a été remis.",
        }
        if new_status in msgs:
            create_notification(
                m.user,
                "Statut mis à jour",
                msgs[new_status],
                category='measurement'
            )
        messages.success(request, f"Statut mis à jour : {m.get_status_display()}")
    return redirect('tech_dashboard')

# ─── RETOUR D'EFFETS (Agent) ──────────────────────────────
@login_required
def enregistrer_retour(request):
    if request.user.role != 'agent':
        return redirect('home')
    if request.method == 'POST':
        form = RetourEffetForm(request.POST)
        if form.is_valid():
            r = form.save(commit=False)
            r.agent = request.user
            r.created_by = request.user
            r.save()
            from accounts.models import User
            from notifications.models import Notification
            admins = User.objects.filter(role='admin')
            msg = f"{request.user.get_full_name()} a retourné {r.quantite} x {r.get_type_equipement_display()} pour {r.get_motif_display()}."
            notifications = [Notification(user=a, title='Retour d\'effet déclaré', message=msg, category='stock') for a in admins]
            if notifications:
                Notification.objects.bulk_create(notifications)
            messages.success(request, "Retour enregistré avec succès.")
            return redirect('agent_dashboard')
    else:
        form = RetourEffetForm()
    return render(request, 'agent/retour.html', {'form': form})


# ─── HISTORIQUE PAR AGENT (Admin) ─────────────────────────
@role_required(['admin'])
def historique_agents(request):
    from django.db.models import Prefetch
    from accounts.models import User

    agent_filter = request.GET.get('agent', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')

    agents = User.objects.filter(role='agent').order_by('nom')
    if agent_filter:
        agents = agents.filter(pk=agent_filter)

    m_qs = Measurement.objects.select_related('rempli_par')
    r_qs = RetourEffet.objects.all()
    if date_debut:
        from datetime import datetime, time
        dt = datetime.combine(datetime.strptime(date_debut, '%Y-%m-%d').date(), time.min)
        m_qs = m_qs.filter(created_at__gte=dt)
        r_qs = r_qs.filter(created_at__gte=dt)
    if date_fin:
        from datetime import datetime, time
        dt = datetime.combine(datetime.strptime(date_fin, '%Y-%m-%d').date(), time.max)
        m_qs = m_qs.filter(created_at__lte=dt)
        r_qs = r_qs.filter(created_at__lte=dt)

    agents = agents.prefetch_related(
        Prefetch('measurements', queryset=m_qs),
        Prefetch('retours', queryset=r_qs),
    )

    historiques = [{'agent': a, 'measurements': a.measurements.all(), 'retours': a.retours.all()} for a in agents]

    return render(request, 'admin_panel/historique.html', {
        'historiques': historiques,
        'agents': agents,
        'agent_filter': agent_filter,
        'date_debut': date_debut,
        'date_fin': date_fin,
    })


# ─── EXPORT EXCEL ─────────────────────────────────────────
@role_required(['admin'])
def export_historique_excel(request):
    from accounts.models import User
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historique par agent"

    # ── Colors ──
    navy_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
    navy_dark_fill = PatternFill(start_color="0F2444", end_color="0F2444", fill_type="solid")
    gold_fill = PatternFill(start_color="C8A84B", end_color="C8A84B", fill_type="solid")
    light_gray_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    header_demande_fill = PatternFill(start_color="1A3A6B", end_color="1A3A6B", fill_type="solid")
    header_retour_fill = PatternFill(start_color="991B1B", end_color="991B1B", fill_type="solid")
    white_font = Font(bold=True, color="FFFFFF", size=9)
    white_font_lg = Font(bold=True, color="FFFFFF", size=14)
    white_font_sm = Font(bold=False, color="B0C4DE", size=8)
    title_font = Font(bold=True, color="1A3A6B", size=12)
    subtitle_font = Font(bold=False, color="64748B", size=8)
    agent_font = Font(bold=True, color="1A3A6B", size=9)
    data_font = Font(size=8.5)
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1'),
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(vertical="center")

    # ── Header block (rows 1-3) ──
    # Banner image (ADII header)
    from openpyxl.drawing.image import Image
    from pathlib import Path
    banner_path = Path(__file__).resolve().parent.parent / 'static' / 'img' / 'adii_header_white.png'
    if banner_path.exists():
        ws.row_dimensions[1].height = 80
        img = Image(str(banner_path))
        img.anchor = 'A1'
        img.width = 290
        img.height = 80
        ws.add_image(img)

    # Row 2: report info
    ws.merge_cells('A3:K3')
    info_cell = ws.cell(row=3, column=1)
    now_str = datetime.now().strftime('%d/%m/%Y %H:%M')
    agent_filter_str = request.GET.get('agent', '')
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    filter_text = "Tous les agents"
    if agent_filter_str:
        filter_agent = User.objects.filter(pk=agent_filter_str).first()
        if filter_agent:
            filter_text = filter_agent.get_full_name()
    info_cell.value = f"Généré le {now_str} — Filtre : {filter_text}"
    if date_debut or date_fin:
        info_cell.value += f" — Période : {date_debut or '...'} au {date_fin or '...'}"
    info_cell.font = subtitle_font
    info_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[3].height = 22

    # Row 4: empty spacer
    ws.row_dimensions[4].height = 6

    # ── Data header row (row 5) ──
    data_start_row = 5
    demande_headers = ["Agent", "Matricule", "Service",
                       "Date demande", "Type équipement", "Statut"]
    retour_headers = ["Date retour", "Type équipement", "Quantité", "Motif retour", "Notes"]

    # Demande headers (blue)
    for col, h in enumerate(demande_headers, 1):
        cell = ws.cell(row=data_start_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=8)
        cell.fill = header_demande_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Retour headers (red)
    for col, h in enumerate(retour_headers, 7):
        cell = ws.cell(row=data_start_row, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF", size=8)
        cell.fill = header_retour_fill
        cell.alignment = center_align
        cell.border = thin_border

    ws.row_dimensions[data_start_row].height = 22

    # ── Data rows ──
    row = data_start_row + 1
    agent_filter_val = request.GET.get('agent', '')
    date_debut_val = request.GET.get('date_debut', '')
    date_fin_val = request.GET.get('date_fin', '')

    from django.db.models import Prefetch
    agents_qs = User.objects.filter(role='agent').order_by('nom')
    if agent_filter_val:
        agents_qs = agents_qs.filter(pk=agent_filter_val)

    m_qs = Measurement.objects.all()
    r_qs = RetourEffet.objects.all()
    if date_debut_val:
        m_qs = m_qs.filter(created_at__date__gte=date_debut_val)
        r_qs = r_qs.filter(created_at__date__gte=date_debut_val)
    if date_fin_val:
        m_qs = m_qs.filter(created_at__date__lte=date_fin_val)
        r_qs = r_qs.filter(created_at__date__lte=date_fin_val)

    agents_qs = agents_qs.prefetch_related(
        Prefetch('measurements', queryset=m_qs),
        Prefetch('retours', queryset=r_qs),
    )

    for agent in agents_qs:
        measurements = list(agent.measurements.all())
        retours = list(agent.retours.all())
        max_rows = max(len(measurements), len(retours), 1)

        for i in range(max_rows):
            m = measurements[i] if i < len(measurements) else None
            r = retours[i] if i < len(retours) else None
            is_even = (row % 2 == 0)

            # Agent info (cols 1-3)
            if i == 0:
                ws.cell(row=row, column=1, value=agent.get_full_name()).font = agent_font
                ws.cell(row=row, column=2, value=agent.matricule or '').font = data_font
                ws.cell(row=row, column=3, value=agent.service or '').font = data_font
            else:
                ws.cell(row=row, column=1, value='').font = data_font
                ws.cell(row=row, column=2, value='').font = data_font
                ws.cell(row=row, column=3, value='').font = data_font

            # Demande data (cols 4-6)
            if m:
                ws.cell(row=row, column=4, value=m.created_at.strftime('%d/%m/%Y')).font = data_font
                ws.cell(row=row, column=5, value=m.get_type_equipement_display()).font = data_font
                ws.cell(row=row, column=6, value=m.get_status_display()).font = Font(size=8.5, bold=True,
                    color='065F46' if m.status == 'livre' else
                          '065F46' if m.status == 'pret' else
                          '5B21B6' if m.status == 'en_production' else
                          '1E40AF' if m.status == 'valide' else
                          '991B1B' if m.status == 'refuse' else '92400E')
            else:
                for c in [4, 5, 6]:
                    ws.cell(row=row, column=c, value='').font = data_font

            # Retour data (cols 7-11)
            if r:
                ws.cell(row=row, column=7, value=r.created_at.strftime('%d/%m/%Y')).font = data_font
                ws.cell(row=row, column=8, value=r.get_type_equipement_display()).font = data_font
                ws.cell(row=row, column=9, value=r.quantite).font = Font(size=8.5, bold=True)
                ws.cell(row=row, column=10, value=r.get_motif_display()).font = Font(size=8.5,
                    color='991B1B' if r.motif == 'destruction' else
                          '92400E' if r.motif == 'perte' else
                          '3730A3' if r.motif == 'usure' else '475569')
                ws.cell(row=row, column=11, value=r.notes or '—').font = data_font
            else:
                for c in [7, 8, 9, 10, 11]:
                    ws.cell(row=row, column=c, value='').font = data_font

            # Borders and row shading
            bg = light_gray_fill if is_even else PatternFill(fill_type=None)
            for col in range(1, 12):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = left_align
                if not (col == 6 and m):
                    cell.fill = bg

            ws.row_dimensions[row].height = 18
            row += 1

    # ── Column widths ──
    col_widths = [22, 12, 14, 14, 18, 14, 14, 18, 10, 14, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Freeze panes ──
    ws.freeze_panes = f'A{data_start_row + 1}'

    # ── Print setup ──
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    # ── Response ──
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = "historique_agents.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ─── EXPORT PDF ───────────────────────────────────────────
def _draw_header(canvas, doc):
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.pagesizes import A4, landscape
    w, h = landscape(A4)
    canvas.saveState()
    canvas.setFillColor(colors.HexColor('#1A3A6B'))
    canvas.roundRect(0.8*cm, h - 1.4*cm, w - 1.6*cm, 0.9*cm, 4, fill=1, stroke=0)
    canvas.setFillColor(colors.white)
    canvas.setFont('Helvetica-Bold', 11)
    canvas.drawString(1.5*cm, h - 1.1*cm, 'ADII')
    canvas.setFont('Helvetica', 6.5)
    canvas.drawString(1.5*cm, h - 1.35*cm, 'Administration des Douanes et Impots Indirects')
    canvas.setFont('Helvetica', 6)
    canvas.setFillColor(colors.HexColor('#94A3B8'))
    canvas.drawCentredString(w/2, 0.5*cm, f'ADII — Page {doc.page}')
    canvas.restoreState()


def _render_pdf_reportlab(historiques, agents, agent_filter, date_debut, date_fin):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, PageBreak)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        title="Historique par agent",
        topMargin=2.2*cm, bottomMargin=1.2*cm,
        leftMargin=1.2*cm, rightMargin=1.2*cm,
    )
    styles = getSampleStyleSheet()

    title_style = ParagraphStyle(
        'ReportTitle', parent=styles['Title'],
        fontSize=12, textColor=colors.HexColor('#1e293b'),
        spaceAfter=2, spaceBefore=2,
    )
    subtitle_style = ParagraphStyle(
        'ReportSub', parent=styles['Normal'],
        fontSize=7, textColor=colors.HexColor('#64748B'),
        spaceAfter=12,
    )
    agent_heading = ParagraphStyle(
        'AgentHead', parent=styles['Heading2'],
        fontSize=9, textColor=colors.HexColor('#1e293b'),
        spaceBefore=4, spaceAfter=2,
    )
    section_label = ParagraphStyle(
        'SectionLbl', parent=styles['Normal'],
        fontSize=8, textColor=colors.HexColor('#475569'),
        spaceBefore=4, spaceAfter=2,
    )

    elements = []

    from datetime import datetime
    elements.append(Paragraph("Historique par agent", title_style))
    filter_parts = [f"Genere le {datetime.now().strftime('%d/%m/%Y %H:%M')}"]
    filter_parts.append('Agent specifique' if agent_filter else 'Tous les agents')
    if date_debut or date_fin:
        filter_parts.append(f"Du {date_debut or '...'} au {date_fin or '...'}")
    elements.append(Paragraph(' • '.join(filter_parts), subtitle_style))

    for h in historiques:
        agent = h['agent']
        measurements = list(h['measurements'])
        retours = list(h['retours'])
        if not measurements and not retours:
            continue

        agent_info = f"{agent.get_full_name()} — {agent.matricule or 'N/A'} — {agent.service or 'N/A'}"
        badge_text = f"  [{len(measurements)} demande(s) | {len(retours)} retour(s)]"
        elements.append(Paragraph(agent_info + badge_text, agent_heading))

        if measurements:
            elements.append(Paragraph("Demandes d'equipement", section_label))
            data = [["Date", "Type d'equipement", "Statut", "Rempli par"]]
            for m in measurements:
                rempli = m.rempli_par.get_full_name() if m.rempli_par else '—'
                data.append([
                    m.created_at.strftime('%d/%m/%Y'),
                    m.get_type_equipement_display(),
                    m.get_status_display(),
                    rempli,
                ])
            table = Table(data, colWidths=[70, 130, 90, 100])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#475569')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 3))

        if retours:
            elements.append(Paragraph("Retours d'effets", section_label))
            data = [["Date", "Type d'equipement", "Qte", "Motif", "Notes"]]
            for r in retours:
                data.append([
                    r.created_at.strftime('%d/%m/%Y'),
                    r.get_type_equipement_display(),
                    str(r.quantite),
                    r.get_motif_display(),
                    r.notes or '—',
                ])
            table = Table(data, colWidths=[70, 130, 30, 70, 90])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f1f5f9')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#475569')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#e2e8f0')),
                ('TOPPADDING', (0, 0), (-1, -1), 2),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ]))
            elements.append(table)
            elements.append(Spacer(1, 3))

        elements.append(Spacer(1, 4))

    doc.build(elements, onFirstPage=_draw_header, onLaterPages=_draw_header)
    buffer.seek(0)
    return buffer


@role_required(['admin'])
def export_historique_pdf(request):
    from django.db.models import Prefetch
    from accounts.models import User
    date_debut = request.GET.get('date_debut', '')
    date_fin = request.GET.get('date_fin', '')
    agent_filter = request.GET.get('agent', '')

    agents_qs = User.objects.filter(role='agent').order_by('nom')
    if agent_filter:
        agents_qs = agents_qs.filter(pk=agent_filter)

    m_qs = Measurement.objects.select_related('rempli_par')
    r_qs = RetourEffet.objects.all()
    if date_debut:
        m_qs = m_qs.filter(created_at__date__gte=date_debut)
        r_qs = r_qs.filter(created_at__date__gte=date_debut)
    if date_fin:
        m_qs = m_qs.filter(created_at__date__lte=date_fin)
        r_qs = r_qs.filter(created_at__date__lte=date_fin)

    agents_qs = agents_qs.prefetch_related(
        Prefetch('measurements', queryset=m_qs),
        Prefetch('retours', queryset=r_qs),
    )

    historiques = [{'agent': a, 'measurements': a.measurements.all(), 'retours': a.retours.all()} for a in agents_qs]

    try:
        import weasyprint
        html = render_to_string('admin_panel/historique_pdf.html', {
            'historiques': historiques,
            'agent_filter': agent_filter,
            'date_debut': date_debut,
            'date_fin': date_fin,
        })
        from pathlib import Path
        pdf = weasyprint.HTML(string=html, base_url=Path(__file__).resolve().parent.parent).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="historique_agents.pdf"'
        return response
    except (ImportError, OSError):
        buffer = _render_pdf_reportlab(historiques, agents_qs, agent_filter, date_debut, date_fin)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="historique_agents.pdf"'
        return response


@login_required
def measurement_detail(request, pk):
    m = get_object_or_404(Measurement.objects.select_related('user', 'rempli_par'), pk=pk)
    if request.user.role == 'agent' and m.user != request.user:
        return redirect('home')
    return render(request, 'measurement_detail.html', {'m': m})


@login_required
@role_required(['admin'])
def reception_pdf(request):
    ids = request.GET.get('ids', '')
    if not ids:
        return HttpResponseBadRequest("Missing ids parameter")

    pk_list = [int(x) for x in ids.split(',') if x.isdigit()]
    fiches = Measurement.objects.filter(pk__in=pk_list, status__in=['termine', 'livre']).select_related('user')

    if not fiches.exists():
        return HttpResponseBadRequest("No matching records found")

    from datetime import datetime
    try:
        import weasyprint
        from pathlib import Path
        html = render_to_string('admin_panel/reception_pdf.html', {
            'fiches': fiches,
            'date_remise': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'responsable': request.user.get_full_name() or request.user.username,
        })
        pdf = weasyprint.HTML(string=html, base_url=Path(__file__).resolve().parent.parent).write_pdf()
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="bon_reception.pdf"'
        return response
    except (ImportError, OSError):
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import cm
        from reportlab.lib import colors
        from reportlab.pdfgen import canvas as rl_canvas

        buffer = io.BytesIO()
        c = rl_canvas.Canvas(buffer, pagesize=A4)
        w, h = A4

        c.setFillColor(colors.HexColor('#1e293b'))
        c.setFont('Times-Bold', 14)
        c.drawCentredString(w/2, h - 1.8*cm, 'ADII')
        c.setFont('Times-Roman', 8)
        c.drawCentredString(w/2, h - 2.2*cm, 'Administration des Douanes et Impôts Indirects')

        c.setFont('Times-Bold', 16)
        c.drawCentredString(w/2, h - 3*cm, 'Bon de réception')

        c.setFont('Times-Roman', 9)
        c.drawCentredString(w/2, h - 3.8*cm, 'Remise d\'équipement d\'habillement')

        c.setStrokeColor(colors.HexColor('#cbd5e1'))
        c.line(2*cm, h - 4.5*cm, w - 2*cm, h - 4.5*cm)

        c.setFont('Times-Roman', 9)
        c.drawString(2*cm, h - 5.2*cm, f'Date: {datetime.now().strftime("%d/%m/%Y %H:%M")}')
        c.drawString(2*cm, h - 5.7*cm, f'Responsable: {request.user.get_full_name() or request.user.username}')

        c.line(2*cm, h - 6.2*cm, w - 2*cm, h - 6.2*cm)

        y = h - 6.8*cm
        c.setFont('Times-Bold', 9)
        c.drawString(2*cm, y, 'Agent')
        c.drawString(6*cm, y, 'Matricule')
        c.drawString(10*cm, y, 'Équipement')
        y -= 0.5*cm
        c.setStrokeColor(colors.HexColor('#e2e8f0'))
        c.line(2*cm, y, w - 2*cm, y)
        y -= 0.3*cm

        c.setFont('Times-Roman', 9)
        for f in fiches:
            if y < 3*cm:
                c.showPage()
                y = h - 2*cm
                c.setFont('Times-Bold', 9)
                c.drawString(2*cm, y, 'Agent')
                c.drawString(6*cm, y, 'Matricule')
                c.drawString(10*cm, y, 'Équipement')
                y -= 0.5*cm
                c.line(2*cm, y, w - 2*cm, y)
                y -= 0.3*cm
                c.setFont('Times-Roman', 9)
            c.drawString(2*cm, y, f.user.get_full_name())
            c.drawString(6*cm, y, f.user.matricule or '—')
            c.drawString(10*cm, y, f.get_type_equipement_display())
            y -= 0.5*cm

        c.showPage()
        c.save()
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="bon_reception.pdf"'
        return response


@login_required
@role_required(['admin'])
def liste_bons_reception(request):
    from django.db.models import Count

    livrees = Measurement.objects.filter(
        status='livre'
    ).select_related('user', 'rempli_par').order_by('-updated_at')[:200]

    group_ids = request.GET.get('ids', '')
    fiches_selection = None
    if group_ids:
        pk_list = [int(x) for x in group_ids.split(',') if x.isdigit()]
        fiches_selection = livrees.filter(pk__in=pk_list)

    return render(request, 'admin_panel/bons_reception.html', {
        'livrees': livrees,
        'fiches_selection': fiches_selection,
    })